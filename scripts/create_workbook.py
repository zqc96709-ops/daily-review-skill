#!/usr/bin/env python3
"""Create a visual time-tracking and daily-review workbook for solo ecommerce sellers."""

from __future__ import annotations

import argparse
import os
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

for candidate in [
    Path.home() / ".cache/codex-runtimes/codex-primary-runtime/dependencies/python/lib/python3.12/site-packages",
    Path.home() / ".cache/codex-runtimes/codex-primary-runtime/dependencies/python",
]:
    if candidate.exists() and str(candidate) not in sys.path:
        sys.path.insert(0, str(candidate))

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import BarChart, DoughnutChart, LineChart, PieChart, RadarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.marker import Marker
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:
    bundled_python = Path.home() / ".cache/codex-runtimes/codex-primary-runtime/dependencies/python/bin/python3"
    if bundled_python.exists() and Path(sys.executable).resolve() != bundled_python.resolve() and os.environ.get("CBTR_REEXEC") != "1":
        env = os.environ.copy()
        env["CBTR_REEXEC"] = "1"
        os.execve(str(bundled_python), [str(bundled_python), __file__, *sys.argv[1:]], env)
    print("Missing dependency: openpyxl. Install it or run with a Python environment that includes openpyxl.", file=sys.stderr)
    raise SystemExit(2) from exc


CATEGORIES = [
    ("选品/市场调研", "推进型", "找机会、看竞品、筛需求、算利润"),
    ("上架/Listing优化", "推进型", "标题、图片、五点、详情页、关键词"),
    ("广告/数据分析", "推进型", "广告结构、预算、转化、报表分析"),
    ("内容素材", "推进型", "图片、短视频、文案、买家秀素材"),
    ("供应链/采购", "推进型", "找工厂、比价、打样、谈交期"),
    ("客服/售后", "维护型", "客户消息、纠纷、评价处理"),
    ("订单/物流", "维护型", "发货、追踪、异常物流"),
    ("财务/记账", "维护型", "利润核算、账单、税务资料"),
    ("学习/资料整理", "学习型", "课程、文章、方法沉淀"),
    ("沟通/会议", "维护型", "合作方沟通、平台沟通"),
    ("行政/工具/杂事", "杂事", "插件、表格、工具、账号维护"),
    ("休息/生活", "休息", "吃饭、运动、家务、休息"),
    ("干扰/刷信息", "干扰", "无明确目的的信息流、低价值浏览"),
]

WORK_TYPES = ["推进型", "维护型", "学习型", "杂事", "休息", "干扰"]
DISTRACTIONS = ["手机/短视频", "平台消息", "客户消息", "临时想法", "工具问题", "家务/生活", "疲劳拖延", "其他"]

COLORS = {
    "dark": "20343B",
    "teal": "2F8F83",
    "mint": "DFF2ED",
    "gold": "FFF3D4",
    "light": "F8FAFA",
    "text": "253238",
}

FONT_NAME = "PingFang SC"
THIN = Side(style="thin", color="D7E0E2")
MEDIUM = Side(style="medium", color="9FB5BA")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", "-o", default=str(Path.home() / "Desktop" / "跨境电商时间管理复盘模板.xlsx"))
    parser.add_argument("--start-date", default=date.today().isoformat(), help="First date in YYYY-MM-DD format.")
    parser.add_argument("--days", type=int, default=90, help="Number of daily review/summary rows.")
    parser.add_argument("--time-rows", type=int, default=520, help="Number of rows in the time log table.")
    return parser.parse_args()


def title(ws, text: str, subtitle: str | None = None) -> None:
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = text
    cell.fill = PatternFill("solid", fgColor=COLORS["dark"])
    cell.font = Font(name=FONT_NAME, size=20, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32
    if subtitle:
        ws.merge_cells("A2:H2")
        sub = ws["A2"]
        sub.value = subtitle
        sub.fill = PatternFill("solid", fgColor="EAF4F2")
        sub.font = Font(name=FONT_NAME, size=11, color="49636A")
        sub.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[2].height = 30


def style_header(ws, row: int, first_col: int, last_col: int) -> None:
    for col in range(first_col, last_col + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=COLORS["teal"])
        cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)


def borders(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.font.name is None:
                cell.font = Font(name=FONT_NAME, color=COLORS["text"])


def widths(ws, mapping: dict[str, int]) -> None:
    for col, width in mapping.items():
        ws.column_dimensions[col].width = width


def add_table(ws, display_name: str, ref: str, style: str = "TableStyleMedium2") -> None:
    table = Table(displayName=display_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name=style, showFirstColumn=False, showLastColumn=False, showRowStripes=True)
    ws.add_table(table)


def add_validations(time_ws, review_ws, dash_ws, time_rows: int, review_last_row: int) -> None:
    category = DataValidation(type="list", formula1="=设置!$A$5:$A$17", allow_blank=True)
    work_type = DataValidation(type="list", formula1="=设置!$E$5:$E$10", allow_blank=True)
    yes_no = DataValidation(type="list", formula1='"是,否"', allow_blank=True)
    score = DataValidation(type="whole", operator="between", formula1="1", formula2="5", allow_blank=True)
    distraction = DataValidation(type="list", formula1="=设置!$F$5:$F$12", allow_blank=True)
    percent = DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=True)
    dates = DataValidation(type="date", operator="between", formula1="DATE(2020,1,1)", formula2="DATE(2035,12,31)", allow_blank=True)
    times = DataValidation(type="time", operator="between", formula1="TIME(0,0,0)", formula2="TIME(23,59,59)", allow_blank=True)

    for dv in [category, work_type, yes_no, score, distraction, dates, times]:
        time_ws.add_data_validation(dv)
    category.add(f"F5:F{time_rows}")
    work_type.add(f"G5:G{time_rows}")
    yes_no.add(f"H5:H{time_rows}")
    score.add(f"K5:L{time_rows}")
    distraction.add(f"M5:M{time_rows}")
    dates.add(f"A5:A{time_rows}")
    times.add(f"C5:D{time_rows}")

    for dv in [percent, score, distraction]:
        review_ws.add_data_validation(dv)
    percent.add(f"C5:C{review_last_row}")
    score.add(f"I5:K{review_last_row}")
    distraction.add(f"F5:F{review_last_row}")

    selector = DataValidation(type="date", operator="between", formula1="DATE(2020,1,1)", formula2="DATE(2035,12,31)", allow_blank=False)
    dash_ws.add_data_validation(selector)
    selector.add("B3")


def build_workbook(start: date, days: int, time_rows_count: int, output: Path) -> None:
    wb = Workbook()
    instructions = wb.active
    instructions.title = "使用说明"
    settings = wb.create_sheet("设置")
    time_ws = wb.create_sheet("时间记录")
    review_ws = wb.create_sheet("每日复盘")
    summary_ws = wb.create_sheet("日汇总")
    weekly_ws = wb.create_sheet("周复盘")
    dash_ws = wb.create_sheet("可视化仪表盘")

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    build_instructions(instructions)
    build_settings(settings)
    build_time_log(time_ws, time_rows_count)
    build_daily_review(review_ws, start, days)
    build_daily_summary(summary_ws, days, time_rows_count)
    build_weekly_review(weekly_ws, start, days)
    build_dashboard(dash_ws, start, time_rows_count)
    add_validations(time_ws, review_ws, dash_ws, time_rows_count, 4 + days)

    wb._sheets = [instructions, time_ws, review_ws, dash_ws, summary_ws, weekly_ws, settings]
    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    check = load_workbook(output, data_only=False)
    if len(check["可视化仪表盘"]._charts) < 5:
        raise RuntimeError("Workbook was created, but dashboard charts are missing.")


def build_instructions(ws) -> None:
    title(ws, "跨境电商个人卖家时间管理与复盘模板", "目标：看清时间花在哪、每天是否推进业务主线、哪些事情正在吞噬注意力。")
    ws["A4"] = "每天怎么用"
    ws["A4"].font = Font(name=FONT_NAME, size=14, bold=True, color=COLORS["dark"])
    steps = [
        "白天：在“时间记录”里按时间段填实际做了什么。建议 30-120 分钟一个时间块，不必精确到分钟。",
        "晚上：在“每日复盘”里填今日主线、主线完成度、三个产出、干扰源、明日主线。",
        "看图：打开“可视化仪表盘”，选择日期，就能看到当天时间占比、推进/维护比例、近 14 天趋势和复盘评分。",
        "每周：在“周复盘”里写本周战役，总结是否真正推进了选品、上新、广告、供应链等关键动作。",
    ]
    for row, text in enumerate(steps, start=5):
        ws[f"A{row}"] = row - 4
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=COLORS["teal"])
        ws[f"A{row}"].font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{row}"] = text
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row].height = 32
    ws["A11"] = "推荐判断标准"
    ws["A11"].font = Font(name=FONT_NAME, size=14, bold=True, color=COLORS["dark"])
    rules = [
        "推进型时间：选品、上新、Listing 优化、广告测试、内容素材、供应链谈判。这些通常直接影响增长。",
        "维护型时间：客服、售后、订单、物流、通知、沟通。这些必须做，但不应该全天打断你。",
        "每天只设一个主线任务。只要主线没有推进，就算很忙，也要在复盘里如实记录。",
        "复盘可视化来自量化字段：主线完成度、专注评分、能量评分、满意度，以及干扰时间。",
    ]
    for row, text in enumerate(rules, start=12):
        ws[f"A{row}"] = "•"
        ws[f"B{row}"] = text
        ws[f"B{row}"].alignment = Alignment(wrap_text=True)
    widths(ws, {"A": 8, "B": 100})


def build_settings(ws) -> None:
    title(ws, "设置", "这里维护下拉选项。后续如果你的业务流程变化，可以在这里增删类别。")
    for col, header in enumerate(["时间类别", "默认工作类型", "说明"], start=1):
        ws.cell(4, col).value = header
    style_header(ws, 4, 1, 3)
    for row, values in enumerate(CATEGORIES, start=5):
        for col, value in enumerate(values, start=1):
            ws.cell(row, col).value = value
    borders(ws, 5, 4 + len(CATEGORIES), 1, 3)

    ws["E4"] = "工作类型"
    ws["F4"] = "干扰源"
    style_header(ws, 4, 5, 6)
    for row, value in enumerate(WORK_TYPES, start=5):
        ws.cell(row, 5).value = value
    for row, value in enumerate(DISTRACTIONS, start=5):
        ws.cell(row, 6).value = value
    borders(ws, 5, 12, 5, 6)
    widths(ws, {"A": 20, "B": 14, "C": 42, "D": 4, "E": 14, "F": 18})
    ws.freeze_panes = "A5"


def build_time_log(ws, time_rows: int) -> None:
    title(ws, "时间记录", "按时间块记录当天实际投入。类别和工作类型会用于自动生成图表。")
    headers = ["日期", "星期", "开始时间", "结束时间", "时长(小时)", "类别", "工作类型", "是否主线", "具体事项", "产出/结果", "精力(1-5)", "专注(1-5)", "干扰源", "备注"]
    for col, header in enumerate(headers, start=1):
        ws.cell(4, col).value = header
    style_header(ws, 4, 1, len(headers))
    for row in range(5, time_rows + 1):
        ws.cell(row, 2).value = f'=IF(A{row}="","",TEXT(A{row},"aaa"))'
        ws.cell(row, 5).value = f'=IF(OR(C{row}="",D{row}=""),"",MOD(D{row}-C{row},1)*24)'
        ws.cell(row, 7).value = f'=IF(F{row}="","",IFERROR(VLOOKUP(F{row},设置!$A$5:$B$17,2,FALSE),""))'
        ws.cell(row, 1).number_format = "yyyy-mm-dd"
        ws.cell(row, 3).number_format = "hh:mm"
        ws.cell(row, 4).number_format = "hh:mm"
        ws.cell(row, 5).number_format = "0.00"
    borders(ws, 5, time_rows, 1, len(headers))
    widths(ws, {"A": 13, "B": 8, "C": 10, "D": 10, "E": 11, "F": 18, "G": 12, "H": 10, "I": 34, "J": 34, "K": 10, "L": 10, "M": 16, "N": 24})
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:N{time_rows}"
    add_table(ws, "TimeLog", f"A4:N{time_rows}")
    ws.conditional_formatting.add(f"E5:E{time_rows}", ColorScaleRule(start_type="min", start_color="FFFFFF", mid_type="percentile", mid_value=50, mid_color="DFF2ED", end_type="max", end_color="2F8F83"))
    ws.conditional_formatting.add(f"L5:L{time_rows}", ColorScaleRule(start_type="num", start_value=1, start_color="FBE4E6", mid_type="num", mid_value=3, mid_color="FFF3D4", end_type="num", end_value=5, end_color="DFF2ED"))


def build_daily_review(ws, start: date, days: int) -> None:
    title(ws, "每日复盘", "晚上 10 分钟填写。文字帮助你思考，数字会进入复盘可视化。")
    headers = ["日期", "今日主线任务", "主线完成度", "今日3个产出", "最大时间黑洞", "最大干扰源", "明日主线任务", "明日第一步动作", "专注评分(1-5)", "能量评分(1-5)", "满意度(1-5)", "一句话结论"]
    for col, header in enumerate(headers, start=1):
        ws.cell(4, col).value = header
    style_header(ws, 4, 1, len(headers))
    for offset, row in enumerate(range(5, 5 + days)):
        ws.cell(row, 1).value = start + timedelta(days=offset)
        ws.cell(row, 1).number_format = "yyyy-mm-dd"
        ws.cell(row, 3).number_format = "0%"
        ws.row_dimensions[row].height = 42
    borders(ws, 5, 4 + days, 1, len(headers))
    widths(ws, {"A": 13, "B": 28, "C": 13, "D": 38, "E": 24, "F": 16, "G": 28, "H": 28, "I": 13, "J": 13, "K": 13, "L": 38})
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:L{4 + days}"
    add_table(ws, "DailyReview", f"A4:L{4 + days}", "TableStyleMedium4")
    ws.conditional_formatting.add(f"C5:C{4 + days}", ColorScaleRule(start_type="num", start_value=0, start_color="FBE4E6", mid_type="num", mid_value=0.5, mid_color="FFF3D4", end_type="num", end_value=1, end_color="DFF2ED"))
    ws.conditional_formatting.add(f"I5:K{4 + days}", ColorScaleRule(start_type="num", start_value=1, start_color="FBE4E6", mid_type="num", mid_value=3, mid_color="FFF3D4", end_type="num", end_value=5, end_color="DFF2ED"))


def build_daily_summary(ws, days: int, time_rows: int) -> None:
    title(ws, "日汇总", "这里自动汇总“时间记录”和“每日复盘”。通常不用手动改。")
    headers = ["日期"] + [item[0] for item in CATEGORIES] + ["总小时", "推进型小时", "维护型小时", "学习型小时", "杂事小时", "休息小时", "干扰小时", "主线小时", "推进型占比", "维护型占比", "平均专注", "平均精力", "复盘主线完成度", "复盘满意度"]
    for col, header in enumerate(headers, start=1):
        ws.cell(4, col).value = header
    style_header(ws, 4, 1, len(headers))
    date_range = f"时间记录!$A$5:$A${time_rows}"
    hour_range = f"时间记录!$E$5:$E${time_rows}"
    cat_range = f"时间记录!$F$5:$F${time_rows}"
    type_range = f"时间记录!$G$5:$G${time_rows}"
    main_range = f"时间记录!$H$5:$H${time_rows}"
    energy_range = f"时间记录!$K$5:$K${time_rows}"
    focus_range = f"时间记录!$L$5:$L${time_rows}"
    base = 2 + len(CATEGORIES)
    for row in range(5, 5 + days):
        ws.cell(row, 1).value = f"=每日复盘!A{row}"
        ws.cell(row, 1).number_format = "yyyy-mm-dd"
        for col in range(2, 2 + len(CATEGORIES)):
            letter = get_column_letter(col)
            ws.cell(row, col).value = f"=SUMIFS({hour_range},{date_range},$A{row},{cat_range},{letter}$4)"
            ws.cell(row, col).number_format = "0.00"
        ws.cell(row, base).value = f"=SUM(B{row}:N{row})"
        ws.cell(row, base + 1).value = f'=SUMIFS({hour_range},{date_range},$A{row},{type_range},"推进型")'
        ws.cell(row, base + 2).value = f'=SUMIFS({hour_range},{date_range},$A{row},{type_range},"维护型")'
        ws.cell(row, base + 3).value = f'=SUMIFS({hour_range},{date_range},$A{row},{type_range},"学习型")'
        ws.cell(row, base + 4).value = f'=SUMIFS({hour_range},{date_range},$A{row},{type_range},"杂事")'
        ws.cell(row, base + 5).value = f'=SUMIFS({hour_range},{date_range},$A{row},{type_range},"休息")'
        ws.cell(row, base + 6).value = f'=SUMIFS({hour_range},{date_range},$A{row},{type_range},"干扰")'
        ws.cell(row, base + 7).value = f'=SUMIFS({hour_range},{date_range},$A{row},{main_range},"是")'
        ws.cell(row, base + 8).value = f'=IF({get_column_letter(base)}{row}=0,"",{get_column_letter(base + 1)}{row}/{get_column_letter(base)}{row})'
        ws.cell(row, base + 9).value = f'=IF({get_column_letter(base)}{row}=0,"",{get_column_letter(base + 2)}{row}/{get_column_letter(base)}{row})'
        ws.cell(row, base + 10).value = f'=IFERROR(AVERAGEIFS({focus_range},{date_range},$A{row}),"")'
        ws.cell(row, base + 11).value = f'=IFERROR(AVERAGEIFS({energy_range},{date_range},$A{row}),"")'
        ws.cell(row, base + 12).value = f'=IFERROR(VLOOKUP($A{row},每日复盘!$A$5:$K${4 + days},3,FALSE),"")'
        ws.cell(row, base + 13).value = f'=IFERROR(VLOOKUP($A{row},每日复盘!$A$5:$K${4 + days},11,FALSE),"")'
        for col in range(base, base + 8):
            ws.cell(row, col).number_format = "0.00"
        for col in (base + 8, base + 9, base + 12):
            ws.cell(row, col).number_format = "0%"
        for col in (base + 10, base + 11, base + 13):
            ws.cell(row, col).number_format = "0.0"
    borders(ws, 5, 4 + days, 1, len(headers))
    widths(ws, {get_column_letter(i): 14 for i in range(1, len(headers) + 1)})
    ws.column_dimensions["A"].width = 13
    ws.freeze_panes = "B5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{4 + days}"


def build_weekly_review(ws, start: date, days: int) -> None:
    title(ws, "周复盘", "每周看一次：这一周到底推进了哪个战役，而不是只处理了多少杂事。")
    headers = ["周开始", "周结束", "本周战役", "本周胜利成果", "下周要改进", "推进型小时", "维护型小时", "干扰小时", "主线完成均值", "满意度均值"]
    for col, header in enumerate(headers, start=1):
        ws.cell(4, col).value = header
    style_header(ws, 4, 1, len(headers))
    first_monday = start - timedelta(days=start.weekday())
    week_count = max(8, (days + 6) // 7 + 1)
    for index, row in enumerate(range(5, 5 + week_count)):
        week_start = first_monday + timedelta(days=7 * index)
        ws.cell(row, 1).value = week_start
        ws.cell(row, 2).value = week_start + timedelta(days=6)
        ws.cell(row, 6).value = f'=SUMIFS(日汇总!$P$5:$P${4 + days},日汇总!$A$5:$A${4 + days},">="&A{row},日汇总!$A$5:$A${4 + days},"<="&B{row})'
        ws.cell(row, 7).value = f'=SUMIFS(日汇总!$Q$5:$Q${4 + days},日汇总!$A$5:$A${4 + days},">="&A{row},日汇总!$A$5:$A${4 + days},"<="&B{row})'
        ws.cell(row, 8).value = f'=SUMIFS(日汇总!$U$5:$U${4 + days},日汇总!$A$5:$A${4 + days},">="&A{row},日汇总!$A$5:$A${4 + days},"<="&B{row})'
        ws.cell(row, 9).value = f'=IFERROR(AVERAGEIFS(日汇总!$AA$5:$AA${4 + days},日汇总!$A$5:$A${4 + days},">="&A{row},日汇总!$A$5:$A${4 + days},"<="&B{row}),"")'
        ws.cell(row, 10).value = f'=IFERROR(AVERAGEIFS(日汇总!$AB$5:$AB${4 + days},日汇总!$A$5:$A${4 + days},">="&A{row},日汇总!$A$5:$A${4 + days},"<="&B{row}),"")'
        for col in (1, 2):
            ws.cell(row, col).number_format = "yyyy-mm-dd"
        for col in (6, 7, 8):
            ws.cell(row, col).number_format = "0.00"
        ws.cell(row, 9).number_format = "0%"
        ws.cell(row, 10).number_format = "0.0"
        ws.row_dimensions[row].height = 44
    borders(ws, 5, 4 + week_count, 1, len(headers))
    widths(ws, {"A": 13, "B": 13, "C": 28, "D": 38, "E": 38, "F": 13, "G": 13, "H": 13, "I": 14, "J": 13})
    ws.freeze_panes = "A5"


def build_dashboard(ws, start: date, time_rows: int) -> None:
    title(ws, "可视化仪表盘", "在 B3 选择日期。图表会显示当天时间占比、工作类型结构、近 14 天趋势，以及复盘是否聚焦。")
    ws["A3"] = "选择日期"
    ws["B3"] = start
    ws["B3"].number_format = "yyyy-mm-dd"
    for cell in ("A3", "B3"):
        ws[cell].fill = PatternFill("solid", fgColor=COLORS["gold"])
        ws[cell].border = Border(top=MEDIUM, left=MEDIUM, right=MEDIUM, bottom=MEDIUM)
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")
    ws["A3"].font = Font(name=FONT_NAME, bold=True, color=COLORS["dark"])

    build_dashboard_tables(ws, time_rows)
    build_dashboard_charts(ws)
    widths(ws, {"A": 18, "B": 12, "C": 3, "D": 14, "E": 12, "F": 3, "G": 22, "H": 12, "I": 3, "J": 12, "K": 12, "L": 12, "M": 12, "N": 12, "O": 12, "P": 12, "Q": 12})
    for row in range(1, 58):
        ws.row_dimensions[row].height = 22
    ws.freeze_panes = "A6"


def build_dashboard_tables(ws, time_rows: int) -> None:
    ws["A5"] = "当天类别小时"
    ws["A6"] = "类别"
    ws["B6"] = "小时"
    style_header(ws, 6, 1, 2)
    for row, (category, _, _) in enumerate(CATEGORIES, start=7):
        ws.cell(row, 1).value = category
        ws.cell(row, 2).value = f'=SUMIFS(时间记录!$E$5:$E${time_rows},时间记录!$A$5:$A${time_rows},$B$3,时间记录!$F$5:$F${time_rows},A{row})'
        ws.cell(row, 2).number_format = "0.00"
    borders(ws, 7, 19, 1, 2)

    ws["D5"] = "当天工作类型"
    ws["D6"] = "类型"
    ws["E6"] = "小时"
    style_header(ws, 6, 4, 5)
    for row, work_type in enumerate(WORK_TYPES, start=7):
        ws.cell(row, 4).value = work_type
        ws.cell(row, 5).value = f'=SUMIFS(时间记录!$E$5:$E${time_rows},时间记录!$A$5:$A${time_rows},$B$3,时间记录!$G$5:$G${time_rows},D{row})'
        ws.cell(row, 5).number_format = "0.00"
    borders(ws, 7, 12, 4, 5)

    ws["G5"] = "当天复盘评分"
    ws["G6"] = "指标"
    ws["H6"] = "分值"
    style_header(ws, 6, 7, 8)
    metrics = [
        ("主线完成度(折算5分)", '=IFERROR(VLOOKUP($B$3,每日复盘!$A$5:$K$200,3,FALSE)*5,"")'),
        ("专注评分", '=IFERROR(VLOOKUP($B$3,每日复盘!$A$5:$K$200,9,FALSE),"")'),
        ("能量评分", '=IFERROR(VLOOKUP($B$3,每日复盘!$A$5:$K$200,10,FALSE),"")'),
        ("满意度", '=IFERROR(VLOOKUP($B$3,每日复盘!$A$5:$K$200,11,FALSE),"")'),
    ]
    for row, (label, formula) in enumerate(metrics, start=7):
        ws.cell(row, 7).value = label
        ws.cell(row, 8).value = formula
        ws.cell(row, 8).number_format = "0.0"
    borders(ws, 7, 10, 7, 8)

    kpis = [
        ("J3", "当天总小时", "K3", "=SUM(B7:B19)", "0.00"),
        ("L3", "推进型小时", "M3", f'=SUMIFS(时间记录!$E$5:$E${time_rows},时间记录!$A$5:$A${time_rows},$B$3,时间记录!$G$5:$G${time_rows},"推进型")', "0.00"),
        ("N3", "维护型小时", "O3", f'=SUMIFS(时间记录!$E$5:$E${time_rows},时间记录!$A$5:$A${time_rows},$B$3,时间记录!$G$5:$G${time_rows},"维护型")', "0.00"),
        ("P3", "干扰小时", "Q3", f'=SUMIFS(时间记录!$E$5:$E${time_rows},时间记录!$A$5:$A${time_rows},$B$3,时间记录!$G$5:$G${time_rows},"干扰")', "0.00"),
        ("J4", "主线小时", "K4", f'=SUMIFS(时间记录!$E$5:$E${time_rows},时间记录!$A$5:$A${time_rows},$B$3,时间记录!$H$5:$H${time_rows},"是")', "0.00"),
        ("L4", "主线完成度", "M4", '=IFERROR(VLOOKUP($B$3,每日复盘!$A$5:$K$200,3,FALSE),"")', "0%"),
        ("N4", "平均专注", "O4", f'=IFERROR(AVERAGEIFS(时间记录!$L$5:$L${time_rows},时间记录!$A$5:$A${time_rows},$B$3),"")', "0.0"),
        ("P4", "满意度", "Q4", '=IFERROR(VLOOKUP($B$3,每日复盘!$A$5:$K$200,11,FALSE),"")', "0.0"),
    ]
    for label_cell, label, value_cell, formula, number_format in kpis:
        ws[label_cell] = label
        ws[value_cell] = formula
        ws[value_cell].number_format = number_format
    for row in range(3, 5):
        for col in range(10, 18):
            cell = ws.cell(row, col)
            cell.fill = PatternFill("solid", fgColor="EAF4F2" if col % 2 else "F5F8F8")
            cell.border = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(name=FONT_NAME, size=9 if col % 2 == 0 else 12, bold=True, color="49636A" if col % 2 == 0 else COLORS["dark"])

    ws["A22"] = "近14天趋势"
    headers = ["日期", "总小时", "推进型", "维护型", "主线", "干扰", "专注", "满意度"]
    for col, header in enumerate(headers, start=1):
        ws.cell(23, col).value = header
    style_header(ws, 23, 1, len(headers))
    for offset, row in enumerate(range(24, 38)):
        ws.cell(row, 1).value = f"=$B$3-13+{offset}"
        ws.cell(row, 1).number_format = "mm-dd"
        ws.cell(row, 2).value = f"=SUMIFS(时间记录!$E$5:$E${time_rows},时间记录!$A$5:$A${time_rows},A{row})"
        ws.cell(row, 3).value = f'=SUMIFS(时间记录!$E$5:$E${time_rows},时间记录!$A$5:$A${time_rows},A{row},时间记录!$G$5:$G${time_rows},"推进型")'
        ws.cell(row, 4).value = f'=SUMIFS(时间记录!$E$5:$E${time_rows},时间记录!$A$5:$A${time_rows},A{row},时间记录!$G$5:$G${time_rows},"维护型")'
        ws.cell(row, 5).value = f'=SUMIFS(时间记录!$E$5:$E${time_rows},时间记录!$A$5:$A${time_rows},A{row},时间记录!$H$5:$H${time_rows},"是")'
        ws.cell(row, 6).value = f'=SUMIFS(时间记录!$E$5:$E${time_rows},时间记录!$A$5:$A${time_rows},A{row},时间记录!$G$5:$G${time_rows},"干扰")'
        ws.cell(row, 7).value = f'=IFERROR(AVERAGEIFS(时间记录!$L$5:$L${time_rows},时间记录!$A$5:$A${time_rows},A{row}),"")'
        ws.cell(row, 8).value = f'=IFERROR(VLOOKUP(A{row},每日复盘!$A$5:$K$200,11,FALSE),"")'
    borders(ws, 24, 37, 1, 8)


def build_dashboard_charts(ws) -> None:
    bar = BarChart()
    bar.type = "bar"
    bar.title = "当天各类别耗时"
    bar.y_axis.title = "类别"
    bar.x_axis.title = "小时"
    bar.add_data(Reference(ws, min_col=2, min_row=6, max_row=19), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=7, max_row=19))
    bar.height = 9
    bar.width = 17
    bar.legend = None
    bar.dataLabels = DataLabelList()
    bar.dataLabels.showVal = True
    ws.add_chart(bar, "J7")

    line = LineChart()
    line.title = "近14天投入趋势"
    line.y_axis.title = "小时/评分"
    line.x_axis.title = "日期"
    line.add_data(Reference(ws, min_col=2, max_col=6, min_row=23, max_row=37), titles_from_data=True)
    line.set_categories(Reference(ws, min_col=1, min_row=24, max_row=37))
    line.height = 9
    line.width = 20
    line.legend.position = "b"
    for series in line.series:
        series.marker = Marker("circle")
    ws.add_chart(line, "J22")

    pie = PieChart()
    pie.title = "当天时间占比"
    pie.add_data(Reference(ws, min_col=2, min_row=6, max_row=19), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=1, min_row=7, max_row=19))
    pie.height = 8
    pie.width = 10
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    ws.add_chart(pie, "A40")

    radar = RadarChart()
    radar.type = "filled"
    radar.title = "当天复盘雷达"
    radar.add_data(Reference(ws, min_col=8, min_row=6, max_row=10), titles_from_data=True)
    radar.set_categories(Reference(ws, min_col=7, min_row=7, max_row=10))
    radar.height = 8
    radar.width = 10
    ws.add_chart(radar, "G40")

    doughnut = DoughnutChart()
    doughnut.title = "推进/维护/干扰结构"
    doughnut.add_data(Reference(ws, min_col=5, min_row=6, max_row=12), titles_from_data=True)
    doughnut.set_categories(Reference(ws, min_col=4, min_row=7, max_row=12))
    doughnut.holeSize = 55
    doughnut.height = 8
    doughnut.width = 10
    doughnut.dataLabels = DataLabelList()
    doughnut.dataLabels.showPercent = True
    ws.add_chart(doughnut, "L40")


def main() -> None:
    args = parse_args()
    try:
        start = datetime.strptime(args.start_date, "%Y-%m-%d").date()
    except ValueError as exc:
        raise SystemExit("--start-date must use YYYY-MM-DD format") from exc
    build_workbook(start=start, days=args.days, time_rows_count=args.time_rows, output=Path(args.output).expanduser())
    print(Path(args.output).expanduser())


if __name__ == "__main__":
    main()
